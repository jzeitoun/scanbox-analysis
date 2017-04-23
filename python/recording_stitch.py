import numpy as np
import scipy.io as sio
import os

from pacu.core.io.scanbox.impl2 import ScanboxIO
from pacu.core.io.scanbox.view.trial_merged_roi import TrialMergedROIView

from openpyxl import Workbook
from openpyxl.styles import NamedStyle, PatternFill, Border, Side, Alignment, Protection, Font

#import paramiko
#from paramiko.client import SSHClient, SFTPClient

class stitched_data(object):
    '''
    Create this object in directory that contains the .io files of interest.
    Argument should be list of tuples with the filename first and the workspace name second.
    Example: stitched_dataset = stitched_data([('Day1_000_000','Workspace_1'),('Day1_000_001',Workspace_1')])
    '''
    def __init__(self,fw_array):
        self.fw_array = fw_array
	self.path = os.getcwd()
        self.io = [ScanboxIO(os.path.join(self.path,fw[0])) for fw in self.fw_array]
        self.condition = self.io[0].condition
        self.workspaces = [workspace for data,fw in zip(self.io, self.fw_array) for workspace in data.condition.workspaces if workspace.name == fw[1]]
        self.rois = self.find_matched_rois() 
        #self.merged_rois = [TrialMergedROIView(roi.id,*self.workspaces) for roi in self.rois[0]] 
        self.merged_rois = [TrialMergedROIView(roi.params.cell_id,*self.workspaces) for roi in self.rois[0]] 
        self.refresh_all()
	#self.roi_dict = {'{}{}'.format('id_',merged_roi.rois[0].id):merged_roi.serialize() for merged_roi in self.merged_rois}
        self.roi_dict = {'{}{}'.format('cell_id_',merged_roi.rois[0].params.cell_id):merged_roi.serialize() for merged_roi in self.merged_rois}
	#self.sftp = self.create_SFTP()
    
    def find_matched_rois(self):
        rois = [workspace.rois for data, fw in zip(self.io, self.fw_array) for workspace in data.condition.workspaces if workspace.name == fw[1]]    
        #id_sets = [[roi.id for roi in roi_list] for roi_list in rois]
        id_sets = [[roi.params.cell_id for roi in roi_list] for roi_list in rois]
        list_lengths = [len(s) for s in id_sets]
        shortest_idx = list_lengths.index(min(list_lengths))
        shortest_set = id_sets.pop(shortest_idx)
        matched_ids = set(shortest_set).intersection(*id_sets)
        #matched_rois = [[roi for roi in roi_list if roi.id in matched_ids] for roi_list in rois]
        matched_rois = [[roi for roi in roi_list if roi.params.cell_id in matched_ids] for roi_list in rois]
        return matched_rois

    def refresh_all(self):
        for merged_roi in self.merged_rois:
	    merged_roi.refresh()

    def sorted_orientation_traces(self, merged_roi):
        sorted_orientation_traces = {str(roi.workspace.name):{} for roi in merged_roi.rois}
        for k in sorted_orientation_traces.keys():
            sorted_orientation_traces[k] = [dict(dtorientationsmean.attributes.items()[i] for i in [8,17]) for roi in merged_roi.rois for dtorientationsmean in roi.dtorientationsmeans if roi.workspace.name == k]
	return sorted_orientation_traces
    	
    def export_mat(self,filename=None):
        merged_dict = {}
	merged_dict['filenames'] = [fw[0][:-3] for fw in self.fw_array]
	merged_dict['workspaces'] = [workspace.name for workspace in self.workspaces]
	merged_dict['rois'] = self.roi_dict 
	for merged_roi in self.merged_rois: 
	   #merged_dict['rois']['{}{}'.format('id_',merged_roi.rois[0].id)]['sorted_dtorientationsmeans'] = self.sorted_orientation_traces(merged_roi)
            merged_dict['rois']['{}{}'.format('cell_id_',merged_roi.rois[0].params.cell_id)]['sorted_dtorientationsmeans'] = self.sorted_orientation_traces(merged_roi)
        if filename == None:
            fname = self.fw_array[0][0][:-3] + '_merged.mat'
        else:
            fname = filename + '.mat'
        # save .mat file
        sio.savemat(fname,{'merged_dict':merged_dict})             
       
        # create excel file
        wb = Workbook()
        ws = wb.active
       
        num_rois = len(self.merged_rois)
        sfreqs = self.condition.attributes['sfrequencies']
        num_sf = len(sfreqs)
        idx_list = range(3,num_rois*num_sf,num_sf)
        font = 'Courier New'

        # format header columns
        ws.merge_cells('A1:A2')
        ws.merge_cells('B1:C1')
        ws.merge_cells('D1:E1')
        ws.merge_cells('R1:S1')

        for top,bottom in zip(ws['J1:Q1'][0],ws['J2:Q2'][0]):
            ws.merge_cells('{}{}:{}{}'.format(top.column,top.row,bottom.column,bottom.row))
        
        header = NamedStyle(name='header')
        header.alignment = Alignment(horizontal='center',
                                        vertical='center',
                                        text_rotation=0,
                                        wrap_text=True,
                                        shrink_to_fit=False,
                                        indent=0)
        
        header.font = Font(name=font,
                                size=10,
                                bold=True,
                                italic=False,
                                vertAlign=None,
                                underline='none',
                                strike=False,
                                color='FF000000')
        
        header.border = Border(top=Side(border_style='medium',
                                        color='FF000000'),
                                    bottom=Side(border_style=None,
                                        color='FF000000')
                                    )

        reg_cell = NamedStyle(name='regular')
        reg_cell.alignment = Alignment(horizontal='center',
                                     vertical='center',
                                     text_rotation=0,
                                     wrap_text=True,
                                     shrink_to_fit=False,
                                     indent=0)

        reg_cell.font = Font(name=font,size=10)

        reg_cell.border = Border(top=Side(border_style='medium',
                                                 color='FF000000'),
                                             bottom=Side(border_style=None,
                                                 color='FF000000')
                                             )

        sig_cell = NamedStyle(name='significant')
        sig_cell.alignment = Alignment(horizontal='center',
                                          vertical='center',
                                          text_rotation=0,
                                          wrap_text=True,
                                          shrink_to_fit=False,
                                          indent=0)
        
        sig_cell.font = Font(name=font,size=10)

        sig_cell.border = Border(top=Side(border_style='medium',
                                                 color='FF000000'),
                                             bottom=Side(border_style=None,
                                                 color='FF000000')
                                             )

        sig_cell.fill = PatternFill(start_color='FFFFFF00',
                                        end_color='FFFFFF00',
                                        fill_type='solid')

        # write column titles
        ws['A1'].value = 'Cell ID'
        ws['A1'].style = header
        ws['B1'].value = 'Anova All'
        ws['B1'].style = header
        ws['D1'].value = 'SF Cutoff Rel33'
        ws['D1'].style = header
        ws['F1'].value = 'SF'
        ws['F1'].style = header
        ws['R1'].value = 'Anova Each'
        ws['R1'].style = header
        ws['R2'].value = 'F'
        ws['R2'].style = header
        ws['S2'].value = 'P'
        ws['S2'].style = header

        for cell,val in zip(ws[2][1:9],['F','P','X','Y','Peak','Pref','Bandwidth','Global\nOPref']):
            cell.value = val
            cell.style = header
            
        for cell,val in zip(ws[1][9:17],['@', 'OSI', 'CV', 'DSI', 'Sigma', 'OPref', 'RMax', 'Residual']):
            cell.value = val
            cell.style = header

        for idx,roi in zip(idx_list,self.merged_rois):
            if any([anovaeach.p < .01 for anovaeach in roi.dtanovaeachs]):
                style = sig_cell
            else:
                style = reg_cell

            for top,bottom in zip(ws['A{}:I{}'.format(idx,idx)][0],ws['A{}:I{}'.format(idx+num_sf-1,idx+num_sf-1)][0]):
                ws.merge_cells('{}{}:{}{}'.format(top.column,top.row,bottom.column,bottom.row))

            ws.cell(row=idx,column=1).value = int(roi.rois[0].params.cell_id)
            ws.cell(row=idx,column=1).style = style
            ws.cell(row=idx,column=1).alignment = Alignment(vertical='center',horizontal=None)
            ws.cell(row=idx,column=2).value = roi.dtanovaalls.first.attributes['value']['p']
            ws.cell(row=idx,column=2).style = style
            ws.cell(row=idx,column=3).value = roi.dtanovaalls.first.attributes['value']['f']
            ws.cell(row=idx,column=3).style = style
            ws.cell(row=idx,column=4).value = roi.dtsfreqfits.first.attributes['value']['rc33'].x
            ws.cell(row=idx,column=4).style = style
            ws.cell(row=idx,column=5).value = roi.dtsfreqfits.first.attributes['value']['rc33'].y
            ws.cell(row=idx,column=5).style = style
            ws.cell(row=idx,column=6).value = round(roi.dtsfreqfits.first.attributes['value']['peak'],2)
            ws.cell(row=idx,column=6).style = style
            ws.cell(row=idx,column=7).value = roi.dtsfreqfits.first.attributes['value']['pref']
            ws.cell(row=idx,column=7).style = style
            ws.cell(row=idx,column=8).value = roi.dtsfreqfits.first.attributes['value']['ratio']
            ws.cell(row=idx,column=8).style = style
            ws.cell(row=idx,column=9).value = roi.dtorientationbestprefs.first.attributes['value']  
            ws.cell(row=idx,column=9).style = style
           
            for row in ws.iter_rows(min_col=10, max_col=10, min_row=idx, max_row=idx+num_sf-1):
                for cell,sf in zip(row,sfreqs):
                    cell.value = sf
                    cell.style = style

            for i,row in enumerate(ws.iter_rows(min_col=11, max_col=19, min_row=idx, max_row=idx+num_sf-1)):
                row[0].value = roi.dtorientationsfits[i].attributes['value']['osi']    
                row[0].style = style
                row[1].value = roi.dtorientationsfits[i].attributes['value']['cv']
                row[1].style = style
                row[2].value = roi.dtorientationsfits[i].attributes['value']['dsi']
                row[2].style = style
                row[3].value = roi.dtorientationsfits[i].attributes['value']['sigma']
                row[3].style = style
                row[4].value = roi.dtorientationsfits[i].attributes['value']['o_pref']
                row[4].style = style
                row[5].value = roi.dtorientationsfits[i].attributes['value']['r_max']
                row[5].style = style
                row[6].value = roi.dtorientationsfits[i].attributes['value']['residual']
                row[6].style = style
                row[7].value = roi.dtanovaeachs[i].attributes['f']
                row[7].style = style
                row[8].value = roi.dtanovaeachs[i].attributes['p']
                row[8].style = style
        
        # save excel file
        if filename == None:
            fname = self.fw_array[0][0][:-3] + '_merged.xlsx'
        else:
            fname = filename + '.xlsx'
        wb.save(fname)

    def create_SFTP(self):
        hostname = '128.200.21.98' # glass.bio.uci.edu
        username = 'ht'
        password = 'mge2cortex' 

        ssh = SSHClient()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        ssh.connect(hostname,username,password)
        
        sftp = SFTPClient.from_transport(ssh.get_transport())
        return sftp
